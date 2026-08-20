import hashlib
from decimal import Decimal
from django.conf import settings
from django.utils import timezone
from django.utils.text import slugify
from django.contrib.auth import authenticate, get_user_model
from django.db import transaction
from django.db.models import Count, Q
from rest_framework import permissions, status, viewsets
from rest_framework.authtoken.models import Token
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.renderers import BaseRenderer, JSONRenderer

from django.http import HttpResponse

class BinaryFileRenderer(BaseRenderer):
    media_type = "*/*"
    format = "binary"

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data

from nexus.institutions.services.invoice_pdf import build_invoice_pdf

from nexus.institutions.models import (
    Institution,
    InstitutionStatus,
    AcademicDivision,
    Department,
    AcademicProgram,
    AcademicSession,
    InstitutionalDocument,
    InstitutionalDocumentChunk,
    InstitutionStaff,
    StaffAssignment,
    StaffRoleAtUnit,
    StudentProfile,
    EmbeddingStatus,
    Pathway,
    PathwayMilestone,
    StudentMilestoneSubmission,
    SubmissionStatus,
    DiagnosticAssessment,
    DiagnosticQuestion,
    StudentAssessmentSession,
    AICoachConversation,
    AICoachMessage,
    CounsellingSession,
    CounsellingCaseNote,
    CompanyBankDetail,
    PricingPlan,
    InstitutionInvoice,
    InvoiceStatus,
    DivisionType,
    AwardLevel,
    SiwesPatternChoice,
    SiwesAcademicImpactChoice,
    LearningResource,
    LearningResourceType,
)
from ..services.nigerian_curriculum_blueprint import (
    get_master_blueprints,
    import_blueprint_to_institution,
    generate_hierarchy_csv,
    generate_hierarchy_excel,
)
from ..services.student_roster_blueprint import (
    generate_program_student_excel,
    generate_program_student_csv,
    parse_and_validate_student_roster,
    commit_student_roster_bulk,
)
from ..services.document_parser import DocumentParserService
from ..services.login_otp_service import (
    OTP_LIFETIME_SECONDS,
    issue_login_otp,
    mask_email,
    resend_login_otp,
    verify_login_otp,
)
from ..services.embedding_service import EmbeddingService
from ..services.vector_search_service import VectorSearchService
from ..services.groq_advisor_service import GroqAdvisorService
from ..services.pathway_template_service import PathwayTemplateService
from ..services.student_credential_service import StudentCredentialService
from ..services.psychometric_service import PsychometricService
from ..services.student_ai_coach_service import StudentAICoachService
from .serializers import (
    InstitutionListSerializer,
    InstitutionDetailSerializer,
    AcademicDivisionSerializer,
    DepartmentSerializer,
    AcademicProgramSerializer,
    AcademicSessionSerializer,
    InstitutionalDocumentSerializer,
    InstitutionalDocumentChunkSerializer,
    DocumentSearchQuerySerializer,
    DocumentUploadSerializer,
    LearningResourceSerializer,
    LearningResourceUploadSerializer,
    AIAdvisorQuerySerializer,
    InstitutionStaffSerializer,
    StaffAssignmentSerializer,
    StudentProfileSerializer,
    StudentProfileCreateSerializer,
    PathwayMilestoneSerializer,
    PathwayMilestoneCreateUpdateSerializer,
    PathwayListSerializer,
    PathwayDetailSerializer,
    PathwayCreateSerializer,
    PathwayClonePayloadSerializer,
    PathwayPublishTemplatePayloadSerializer,
    StudentMilestoneSubmissionSerializer,
    StudentSubmissionCreateSerializer,
    StudentSubmissionReviewSerializer,
    StudentCredentialGenerationSerializer,
    StudentEnrollPathwaySerializer,
    StudentDashboardDataSerializer,
    DiagnosticAssessmentListSerializer,
    DiagnosticAssessmentDetailSerializer,
    DiagnosticQuestionSerializer,
    StudentAssessmentSubmitSerializer,
    StudentAssessmentSessionSerializer,
    AICoachMessageSerializer,
    AICoachConversationSerializer,
    AICoachChatPayloadSerializer,
    CounsellingSessionSerializer,
    CounsellingSessionCreateSerializer,
    CounsellingSessionConfirmSerializer,
    CounsellingCaseNoteSerializer,
    StudentDossierSerializer,
    AuthLoginSerializer,
    AuthUserSerializer,
    ForgotPasswordSerializer,
    ResetPasswordSerializer,
    CompanyBankDetailSerializer,
    PricingPlanSerializer,
    InstitutionInvoiceSerializer,
    InvoiceSubmitPaymentSerializer,
    AdminUserSerializer,
    InstitutionRegistrationSerializer,
)

User = get_user_model()


def duplicate_error(message: str) -> Response:
    """Standard 400 response for a uniqueness violation with a clear English message.

    Used instead of letting DB IntegrityErrors surface as 500s: every unique
    field is checked before insert so the caller sees an understandable error.
    """
    return Response({"error": message}, status=status.HTTP_400_BAD_REQUEST)




class InstitutionViewSet(viewsets.ModelViewSet):
    """Full CRUD and governance endpoints for Nigerian Tertiary Institutions."""

    queryset = Institution.objects.all().prefetch_related("divisions", "departments", "programs", "documents", "sessions")
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_serializer_class(self):
        if self.action in ["list"]:
            return InstitutionListSerializer
        return InstitutionDetailSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        regulator = self.request.query_params.get("regulator")
        inst_type = self.request.query_params.get("institution_type")
        state = self.request.query_params.get("state")
        is_partner = self.request.query_params.get("is_founding_partner")
        search = self.request.query_params.get("search")

        if regulator:
            qs = qs.filter(regulator=regulator.upper())
        if inst_type:
            qs = qs.filter(institution_type=inst_type.upper())
        if state:
            qs = qs.filter(state__iexact=state)
        if is_partner is not None:
            qs = qs.filter(is_founding_partner=(is_partner.lower() in ["true", "1", "yes"]))
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(short_name__icontains=search) | Q(slug__icontains=search))
        return qs

    def get_object(self):
        obj = super().get_object()
        if obj.status != InstitutionStatus.ACTIVE and obj.invoices.filter(status=InvoiceStatus.PAID).exists():
            obj.status = InstitutionStatus.ACTIVE
            obj.save(update_fields=["status", "updated_at"])
        return obj

    @action(detail=True, methods=["get"], url_path="tree")
    def tree(self, request, id=None):
        """Returns the full 4-tier hierarchy tree for visual hierarchy tree explorers."""
        institution = self.get_object()
        divisions = institution.divisions.filter(is_active=True).prefetch_related(
            "departments__programs"
        )

        tree_data = {
            "id": str(institution.id),
            "name": institution.name,
            "short_name": institution.short_name,
            "regulator": institution.regulator,
            "institution_type": institution.institution_type,
            "tier_two_term": institution.tier_two_term,
            "divisions_count": divisions.count(),
            "divisions": [
                {
                    "id": str(div.id),
                    "name": div.name,
                    "code": div.code,
                    "division_type": div.division_type,
                    "dean_name": div.dean_name,
                    "departments": [
                        {
                            "id": str(dept.id),
                            "name": dept.name,
                            "code": dept.code,
                            "hod_name": dept.hod_name,
                            "siwes_eligible": dept.siwes_eligible,
                            "programs": [
                                {
                                    "id": str(prog.id),
                                    "name": prog.name,
                                    "program_code": prog.program_code,
                                    "award_level": prog.award_level,
                                    "award_level_display": prog.get_award_level_display(),
                                    "duration_years": prog.duration_years,
                                    "siwes_duration_months": prog.siwes_duration_months,
                                }
                                for prog in dept.programs.filter(is_active=True)
                            ],
                        }
                        for dept in div.departments.filter(is_active=True)
                    ],
                }
                for div in divisions
            ],
        }
        return Response(tree_data)

    @action(detail=True, methods=["get"], url_path="governance-summary")
    def governance_summary(self, request, id=None):
        """Returns executive governance metrics for senate oversight and regulatory readiness."""
        institution = self.get_object()
        divisions_count = institution.divisions.count()
        departments_count = institution.departments.count()
        programs_count = institution.programs.count()
        siwes_eligible_depts = institution.departments.filter(siwes_eligible=True).count()
        siwes_ratio = round((siwes_eligible_depts / departments_count * 100), 1) if departments_count > 0 else 0
        documents_count = institution.documents.count()
        indexed_chunks = InstitutionalDocumentChunk.objects.filter(document__institution=institution).count()
        current_session = institution.sessions.filter(is_current=True).first()

        summary = {
            "institution": {
                "id": str(institution.id),
                "name": institution.name,
                "short_name": institution.short_name,
                "regulator": institution.regulator,
                "tier_two_term": institution.tier_two_term,
                "is_founding_partner": institution.is_founding_partner,
            },
            "hierarchy_metrics": {
                "total_divisions": divisions_count,
                "total_departments": departments_count,
                "total_programs": programs_count,
                "siwes_eligible_departments": siwes_eligible_depts,
                "siwes_eligibility_percentage": siwes_ratio,
            },
            "knowledge_base": {
                "total_documents": documents_count,
                "total_indexed_chunks": indexed_chunks,
                "grounding_status": "Active (pgvector)" if indexed_chunks > 0 else "Pending Ingestion",
            },
            "active_session": {
                "label": current_session.session_label if current_session else "Not set",
                "semester": current_session.get_current_semester_display() if current_session else "N/A",
            },
            "accreditation_readiness": {
                "regulator": institution.get_regulator_display(),
                "taxonomy_aligned": True,
                "curriculum_mapped": programs_count > 0,
                "handbook_ingested": documents_count > 0,
            },
        }
        return Response(summary)

    @action(detail=True, methods=["post"], url_path="search-documents")
    def search_documents(self, request, id=None):
        """Zero-hallucination hybrid semantic & keyword search across ingested institutional documents."""
        institution = self.get_object()
        serializer = DocumentSearchQuerySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        query = serializer.validated_data["query"]
        top_k = serializer.validated_data.get("top_k", 5)
        doc_type = serializer.validated_data.get("doc_type")

        results = VectorSearchService.search_chunks(
            query=query,
            institution_id=str(institution.id),
            doc_type=doc_type or None,
            top_k=top_k,
        )

        return Response({
            "query": query,
            "institution_id": str(institution.id),
            "institution_name": institution.name,
            "total_matches": len(results),
            "results": results,
        })

    @action(detail=True, methods=["post"], url_path="ask-advisor")
    def ask_advisor(self, request, id=None):
        """Zero-hallucination institutional AI advisor synthesizing verified policy answers using Groq Cloud."""
        institution = self.get_object()
        serializer = AIAdvisorQuerySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        division = None
        department = None
        session = None

        if data.get("division"):
            division = AcademicDivision.objects.filter(id=data["division"]).first()
        if data.get("department"):
            department = Department.objects.filter(id=data["department"]).first()
        if data.get("session"):
            session = AcademicSession.objects.filter(id=data["session"]).first()

        response_data = GroqAdvisorService.ask_advisor(
            query=data["query"],
            institution=institution,
            division=division,
            department=department,
            session=session,
            doc_type=data.get("doc_type") or None,
            top_k=data.get("top_k", 5),
        )
        return Response(response_data)

    @action(detail=False, methods=["get"], url_path="hierarchy-blueprints")
    def hierarchy_blueprints(self, request):
        """Returns standard NUC / Nigerian higher education master blueprints."""
        archetype = request.query_params.get("archetype")
        blueprints = get_master_blueprints(archetype=archetype)
        return Response({
            "archetype": archetype or "ALL",
            "total_faculties": len(blueprints),
            "blueprints": blueprints,
        })

    @action(
        detail=False,
        methods=["get"],
        url_path="download-hierarchy-template",
        renderer_classes=[BinaryFileRenderer, JSONRenderer],
    )
    def download_hierarchy_template(self, request):
        """Downloads a professionally styled multi-sheet Excel (.xlsx) or CSV template with SIWES guide and lookups."""
        prepopulate = request.query_params.get("prepopulate", "true").lower() in ["true", "1", "yes"]
        archetype = request.query_params.get("archetype")
        fmt = (
            request.query_params.get("export_format")
            or request.query_params.get("template_format")
            or request.query_params.get("file_format")
            or request.query_params.get("format")
            or "excel"
        ).lower()

        if fmt in ["excel", "xlsx"]:
            excel_bytes = generate_hierarchy_excel(prepopulate=prepopulate, archetype=archetype)
            filename = f"nexus_academic_hierarchy_template_{archetype or 'master'}.xlsx"
            response = HttpResponse(
                excel_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        else:
            csv_data = generate_hierarchy_csv(prepopulate=prepopulate, archetype=archetype)
            filename = f"nexus_academic_hierarchy_template_{archetype or 'master'}.csv"
            response = HttpResponse(csv_data, content_type="text/csv; charset=utf-8")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

    @action(detail=True, methods=["post"], url_path="import-blueprint")
    def import_blueprint(self, request, id=None):
        """Deploys selected master blueprint faculties, departments, and programmes into the institution."""
        institution = self.get_object()
        division_keys = request.data.get("division_keys", ["ALL"])
        stats = import_blueprint_to_institution(institution, division_keys)
        return Response({
            "success": True,
            "message": f"Successfully imported master blueprint hierarchy for {institution.short_name}.",
            "stats": stats,
        })

    @action(detail=True, methods=["post"], url_path="bulk-import-hierarchy")
    def bulk_import_hierarchy(self, request, id=None):
        """
        Parses and batch-provisions 4-Tier Hierarchy (Divisions, Departments, Programmes)
        from uploaded Excel (.xlsx) or CSV file or structured JSON rows with full validation.
        """
        institution = self.get_object()
        rows = request.data.get("rows")
        uploaded_file = request.FILES.get("file")

        if uploaded_file and not rows:
            filename = uploaded_file.name.lower()
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                target_sheet_name = "Academic Hierarchy Data"
                if target_sheet_name in wb.sheetnames:
                    ws = wb[target_sheet_name]
                else:
                    ws = wb.active
                
                rows_iter = list(ws.iter_rows(values_only=True))
                if rows_iter and len(rows_iter) > 1:
                    raw_headers = [str(h or "").strip() for h in rows_iter[0]]
                    rows = []
                    for r in rows_iter[1:]:
                        if not any(r):
                            continue
                        row_dict = {}
                        for idx, h in enumerate(raw_headers):
                            if idx < len(r) and h:
                                val = r[idx]
                                row_dict[h] = str(val).strip() if val is not None else ""
                        if row_dict.get("division_name") or row_dict.get("department_name"):
                            rows.append(row_dict)
            else:
                import io
                import csv
                content = uploaded_file.read().decode("utf-8-sig", errors="ignore")
                reader = csv.DictReader(io.StringIO(content))
                rows = [r for r in reader]

        if not rows or not isinstance(rows, list):
            return Response(
                {"detail": "Please provide either a valid Excel/CSV spreadsheet or a 'rows' array of data."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_divs = 0
        created_depts = 0
        created_progs = 0
        errors = []

        with transaction.atomic():
            for idx, r in enumerate(rows, start=1):
                div_name = (r.get("division_name") or "").strip()
                dept_name = (r.get("department_name") or "").strip()
                prog_name = (r.get("program_name") or "").strip()

                if not div_name or not dept_name:
                    errors.append(f"Row {idx}: Missing division_name or department_name.")
                    continue

                div_code = (r.get("division_code") or "").strip()
                div_type = (r.get("division_type") or "FACULTY").strip().upper()
                dean_name = (r.get("dean_name") or "").strip()
                dean_email = (r.get("dean_email") or "").strip()

                dept_code = (r.get("department_code") or "").strip()
                hod_name = (r.get("hod_name") or "").strip()
                hod_email = (r.get("hod_email") or "").strip()
                siwes_eligible_raw = str(r.get("siwes_eligible", "TRUE")).strip().upper()
                siwes_eligible = siwes_eligible_raw in ["TRUE", "1", "YES", "Y"]

                # 1. Get or Create Division
                division, d_created = AcademicDivision.objects.get_or_create(
                    institution=institution,
                    name=div_name,
                    defaults={
                        "code": div_code,
                        "division_type": div_type if hasattr(DivisionType, div_type) else DivisionType.FACULTY,
                        "dean_name": dean_name,
                        "dean_email": dean_email,
                        "is_active": True,
                    },
                )
                if d_created:
                    created_divs += 1

                # 2. Get or Create Department
                department, dp_created = Department.objects.get_or_create(
                    division=division,
                    name=dept_name,
                    defaults={
                        "institution": institution,
                        "code": dept_code,
                        "hod_name": hod_name,
                        "hod_email": hod_email,
                        "siwes_eligible": siwes_eligible,
                        "is_active": True,
                    },
                )
                if dp_created:
                    created_depts += 1

                # 3. Get or Create Programme if specified
                if prog_name:
                    prog_code = (r.get("program_code") or "").strip()
                    award_raw = (r.get("award_level") or "B_SC").strip().upper()
                    try:
                        dur_years = int(r.get("duration_years") or 4)
                    except (ValueError, TypeError):
                        dur_years = 4

                    try:
                        siwes_months = int(r.get("siwes_duration_months") or 6)
                    except (ValueError, TypeError):
                        siwes_months = 6

                    siwes_pat = (r.get("siwes_pattern") or SiwesPatternChoice.SEM2_300L).strip()

                    program, p_created = AcademicProgram.objects.get_or_create(
                        department=department,
                        name=prog_name,
                        defaults={
                            "institution": institution,
                            "program_code": prog_code,
                            "award_level": award_raw if award_raw in AwardLevel.values else AwardLevel.BSC,
                            "duration_years": dur_years,
                            "siwes_duration_months": siwes_months,
                            "siwes_pattern": siwes_pat if siwes_pat in SiwesPatternChoice.values else SiwesPatternChoice.SEM2_300L,
                            "is_active": True,
                        },
                    )
                    if p_created:
                        created_progs += 1

        return Response({
            "success": True,
            "message": f"Successfully processed {len(rows)} hierarchy records.",
            "stats": {
                "created_divisions": created_divs,
                "created_departments": created_depts,
                "created_programs": created_progs,
                "total_rows_processed": len(rows),
            },
            "errors": errors,
        })

    @action(
        detail=True,
        methods=["get"],
        url_path="download-student-template",
        renderer_classes=[BinaryFileRenderer, JSONRenderer],
    )
    def download_student_template(self, request, id=None):
        """
        Generates and downloads a program-specific Student Cohort Onboarding Spreadsheet (.xlsx / .csv).
        """
        institution = self.get_object()
        prog_id = request.query_params.get("program_id") or request.query_params.get("program")
        if not prog_id:
            return Response(
                {"detail": "Please specify a valid 'program_id' parameter."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            program = AcademicProgram.objects.select_related("department", "department__division").get(
                id=prog_id, institution=institution
            )
        except AcademicProgram.DoesNotExist:
            return Response(
                {"detail": f"Academic Programme '{prog_id}' not found in this institution."},
                status=status.HTTP_404_NOT_FOUND,
            )

        fmt = (request.query_params.get("export_format") or request.query_params.get("format") or "excel").lower()
        clean_prog_code = slugify(program.program_code or program.name).replace("-", "_")

        if fmt in ["excel", "xlsx"]:
            excel_bytes = generate_program_student_excel(program)
            filename = f"nexus_student_template_{clean_prog_code}.xlsx"
            response = HttpResponse(
                excel_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        else:
            csv_data = generate_program_student_csv(program)
            filename = f"nexus_student_template_{clean_prog_code}.csv"
            response = HttpResponse(csv_data, content_type="text/csv; charset=utf-8")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

    @action(detail=True, methods=["post"], url_path="bulk-import-students")
    def bulk_import_students(self, request, id=None):
        """
        Validates and/or bulk-ingests student cohort records for a specific Academic Programme.
        Supports dry_run=true for pre-validation and preview before DB commit.
        """
        institution = self.get_object()
        uploaded_file = request.FILES.get("file")
        fallback_program_id = request.data.get("program_id") or request.data.get("program")
        dry_run = str(request.data.get("dry_run", "")).lower() in ["true", "1", "yes"]
        default_pwd_scheme = request.data.get("default_password_scheme", "matric")

        if not uploaded_file and not request.data.get("rows"):
            return Response(
                {"detail": "Please provide an Excel/CSV file or a 'rows' array."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if uploaded_file:
            parse_result = parse_and_validate_student_roster(
                file_obj=uploaded_file,
                filename=uploaded_file.name,
                institution=institution,
                fallback_program_id=fallback_program_id,
            )
        else:
            rows = request.data.get("rows", [])
            program = None
            if fallback_program_id:
                program = AcademicProgram.objects.filter(id=fallback_program_id, institution=institution).first()
            if not program:
                return Response({"detail": "Please specify a valid program_id."}, status=status.HTTP_400_BAD_REQUEST)
            parse_result = {
                "success": True,
                "program": {
                    "id": str(program.id),
                    "name": program.name,
                    "code": program.program_code,
                    "award_level": program.award_level,
                    "duration_years": program.duration_years,
                    "department_name": program.department.name,
                    "division_name": program.department.division.name,
                },
                "stats": {"total_rows": len(rows), "valid_count": len(rows), "error_count": 0},
                "valid_rows": rows,
                "errors": [],
            }

        if not parse_result.get("success") and not parse_result.get("valid_rows"):
            return Response(parse_result, status=status.HTTP_400_BAD_REQUEST)

        # If dry-run requested, return validation summary without touching DB
        if dry_run:
            return Response({
                "dry_run": True,
                "can_commit": len(parse_result["errors"]) == 0 and len(parse_result["valid_rows"]) > 0,
                **parse_result,
            })

        # Commit to DB
        commit_result = commit_student_roster_bulk(
            institution=institution,
            program_id=parse_result["program"]["id"],
            valid_rows=parse_result["valid_rows"],
            default_password_scheme=default_pwd_scheme,
        )

        return Response({
            "dry_run": False,
            "validation": parse_result,
            **commit_result,
        })


class AcademicDivisionViewSet(viewsets.ModelViewSet):
    """CRUD ViewSet for Academic Divisions (Faculties/Schools/Colleges)."""

    queryset = AcademicDivision.objects.all().select_related("institution")
    serializer_class = AcademicDivisionSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution")
        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        return qs


class DepartmentViewSet(viewsets.ModelViewSet):
    """CRUD ViewSet for Academic Departments."""

    queryset = Department.objects.all().select_related("institution", "division")
    serializer_class = DepartmentSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution")
        div_id = self.request.query_params.get("division")
        siwes = self.request.query_params.get("siwes_eligible")
        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        if div_id:
            qs = qs.filter(division_id=div_id)
        if siwes is not None:
            qs = qs.filter(siwes_eligible=(siwes.lower() in ["true", "1", "yes"]))
        return qs


class AcademicProgramViewSet(viewsets.ModelViewSet):
    """CRUD ViewSet for Academic Degree Programmes and Options."""

    queryset = AcademicProgram.objects.all().select_related("institution", "department")
    serializer_class = AcademicProgramSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution")
        dept_id = self.request.query_params.get("department")
        award = self.request.query_params.get("award_level")
        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        if dept_id:
            qs = qs.filter(department_id=dept_id)
        if award:
            qs = qs.filter(award_level=award.upper())
        return qs


class AcademicSessionViewSet(viewsets.ModelViewSet):
    """CRUD ViewSet for Academic Sessions and Semesters."""

    queryset = AcademicSession.objects.all().select_related("institution")
    serializer_class = AcademicSessionSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution")
        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        return qs

    @action(detail=True, methods=["post"], url_path="set-current")
    def set_current(self, request, id=None):
        """Sets this session as the active/current session for its institution."""
        session = self.get_object()
        semester = request.data.get("current_semester")
        with transaction.atomic():
            AcademicSession.objects.filter(institution=session.institution).update(is_current=False)
            session.is_current = True
            update_fields = ["is_current"]
            if semester in [SemesterChoice.FIRST_SEMESTER, SemesterChoice.SECOND_SEMESTER]:
                session.current_semester = semester
                update_fields.append("current_semester")
            session.save(update_fields=update_fields)
        return Response({"status": "ok", "message": f"{session.session_label} ({session.get_current_semester_display()}) is now current."})


class InstitutionalDocumentViewSet(viewsets.ModelViewSet):
    """CRUD and text chunking endpoints for Institutional Knowledge Base Documents."""

    queryset = InstitutionalDocument.objects.all().select_related("institution", "division", "department").prefetch_related("chunks")
    serializer_class = InstitutionalDocumentSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution")
        doc_type = self.request.query_params.get("doc_type")
        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        if doc_type:
            qs = qs.filter(doc_type=doc_type)
        return qs

    @action(detail=False, methods=["post"], url_path="upload")
    def upload_document(self, request):
        """Uploads a PDF, DOCX, or TXT file, parses and chunks it, generates vector embeddings, and indexes into pgvector."""
        serializer = DocumentUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        try:
            institution = Institution.objects.get(id=data["institution"])
        except Institution.DoesNotExist:
            return Response({"error": "Institution not found"}, status=status.HTTP_404_NOT_FOUND)

        division = AcademicDivision.objects.filter(id=data.get("division")).first() if data.get("division") else None
        department = Department.objects.filter(id=data.get("department")).first() if data.get("department") else None
        session = AcademicSession.objects.filter(id=data.get("session")).first() if data.get("session") else None

        file_obj = request.FILES.get("file")
        if file_obj:
            file_content = file_obj.read()
            raw_text, chunks_data, content_hash = DocumentParserService.parse_and_chunk(file_content, file_obj.name)
        elif data.get("raw_text"):
            raw_text = data["raw_text"]
            file_content = raw_text.encode("utf-8")
            raw_text, chunks_data, content_hash = DocumentParserService.parse_and_chunk(file_content, f"{data['title']}.txt")
        else:
            return Response({"error": "Either file or raw_text must be provided."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            doc = InstitutionalDocument.objects.create(
                institution=institution,
                division=division,
                department=department,
                session=session,
                title=data["title"],
                doc_type=data["doc_type"],
                file=file_obj if file_obj else None,
                file_path=file_obj.name if file_obj else "",
                content_hash=content_hash,
                chunk_count=len(chunks_data),
                embedding_status=EmbeddingStatus.INDEXED,
                raw_text=raw_text,
            )

            # Generate vector embeddings for all chunks
            chunk_texts = [c["content"] for c in chunks_data]
            embeddings = EmbeddingService.embed_texts(chunk_texts)

            created_chunks = []
            for idx, c in enumerate(chunks_data):
                emb = embeddings[idx] if idx < len(embeddings) else None
                created_chunks.append(
                    InstitutionalDocumentChunk(
                        document=doc,
                        chunk_index=c["chunk_index"],
                        page_number=c["page_number"],
                        section_reference=c["section_reference"],
                        content=c["content"],
                        embedding=emb,
                        is_header=c.get("is_header", False),
                    )
                )
            InstitutionalDocumentChunk.objects.bulk_create(created_chunks)

        return Response(
            {
                "status": "ok",
                "message": f"Successfully parsed and indexed {len(created_chunks)} citation-ready chunks into pgvector.",
                "document": InstitutionalDocumentSerializer(doc).data,
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="ingest-text")
    def ingest_text(self, request, id=None):
        """Processes raw text for a document, splitting into citation-ready chunks and generating vector embeddings."""
        doc = self.get_object()
        raw_text = request.data.get("raw_text") or doc.raw_text

        if not raw_text or not raw_text.strip():
            return Response(
                {"error": "No raw_text provided to ingest."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        _, chunks_data, content_hash = DocumentParserService.parse_and_chunk(
            raw_text.encode("utf-8"), f"{doc.title}.txt"
        )

        with transaction.atomic():
            doc.chunks.all().delete()

            chunk_texts = [c["content"] for c in chunks_data]
            embeddings = EmbeddingService.embed_texts(chunk_texts)

            created_chunks = []
            for idx, c in enumerate(chunks_data):
                emb = embeddings[idx] if idx < len(embeddings) else None
                created_chunks.append(
                    InstitutionalDocumentChunk(
                        document=doc,
                        chunk_index=c["chunk_index"],
                        page_number=c["page_number"],
                        section_reference=c["section_reference"],
                        content=c["content"],
                        embedding=emb,
                        is_header=c.get("is_header", False),
                    )
                )

            InstitutionalDocumentChunk.objects.bulk_create(created_chunks)

            doc.content_hash = content_hash
            doc.chunk_count = len(created_chunks)
            doc.embedding_status = EmbeddingStatus.INDEXED
            doc.raw_text = raw_text
            doc.save(update_fields=["content_hash", "chunk_count", "embedding_status", "raw_text", "updated_at"])

        return Response({
            "status": "ok",
            "message": f"Successfully ingested and indexed {len(created_chunks)} chunks into pgvector.",
            "document": InstitutionalDocumentSerializer(doc).data,
        })



class LearningResourceViewSet(viewsets.ModelViewSet):
    """CRUD for institution learning content: YouTube videos and uploaded documents/handouts."""

    queryset = LearningResource.objects.all().select_related("institution", "division", "department", "session")
    serializer_class = LearningResourceSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution")
        resource_type = self.request.query_params.get("resource_type")
        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        if resource_type:
            qs = qs.filter(resource_type=resource_type)
        return qs

    @action(detail=False, methods=["post"], url_path="upload")
    def upload_document(self, request):
        """Uploads a handout/document (PDF, DOCX, PPTX, TXT) and attaches it to a learning resource."""
        serializer = LearningResourceUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        try:
            institution = Institution.objects.get(id=data["institution"])
        except Institution.DoesNotExist:
            return Response({"error": "Institution not found"}, status=status.HTTP_404_NOT_FOUND)

        division = AcademicDivision.objects.filter(id=data.get("division")).first() if data.get("division") else None
        department = Department.objects.filter(id=data.get("department")).first() if data.get("department") else None
        session = AcademicSession.objects.filter(id=data.get("session")).first() if data.get("session") else None

        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"error": "A file is required for document resources."}, status=status.HTTP_400_BAD_REQUEST)

        resource = LearningResource.objects.create(
            institution=institution,
            division=division,
            department=department,
            session=session,
            title=data["title"],
            description=data.get("description", ""),
            resource_type=data["resource_type"],
            file=file_obj,
            file_name=file_obj.name,
            file_size=file_obj.size,
            is_published=True,
        )

        return Response(
            {
                "status": "ok",
                "message": "Learning resource uploaded successfully.",
                "resource": LearningResourceSerializer(resource).data,
            },
            status=status.HTTP_201_CREATED,
        )


class InstitutionStaffViewSet(viewsets.ModelViewSet):
    """CRUD ViewSet for managing institutional staff, deans, and evaluators."""

    queryset = InstitutionStaff.objects.all().select_related("user", "institution", "division", "department")
    serializer_class = InstitutionStaffSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution")
        role = self.request.query_params.get("role")
        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        if role:
            qs = qs.filter(role=role.upper())
        return qs

    def create(self, request, *args, **kwargs):
        email = request.data.get("email")
        name = request.data.get("name", "")
        institution_id = request.data.get("institution")
        role = request.data.get("role", "COUNSELLOR")
        title = request.data.get("title", "")
        division_id = request.data.get("division")
        department_id = request.data.get("department")

        if not email or not institution_id:
            return Response(
                {"error": "Email and institution are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            user, created = User.objects.get_or_create(
                email=email.lower().strip(),
                defaults={"name": name, "is_active": True},
            )
            if created:
                user.set_password("1234!@#$")
                user.save()
            elif name and not user.name:
                user.name = name
                user.save(update_fields=["name"])

            staff, _ = InstitutionStaff.objects.update_or_create(
                user=user,
                institution_id=institution_id,
                defaults={
                    "role": role,
                    "title": title,
                    "division_id": division_id if division_id else None,
                    "department_id": department_id if department_id else None,
                    "is_active": True,
                },
            )

        serializer = self.get_serializer(staff)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class AuthLoginView(APIView):
    """Authenticates institutional user and returns DRF Token + profile.

    Students sign in directly. Staff and platform admins receive a secure
    one-time code by email (OTP) and must verify it before receiving a token.
    """

    permission_classes = [AllowAny]

    def post(self, request):
        serializer = AuthLoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data["email"].lower().strip()
        password = serializer.validated_data["password"]

        user = authenticate(request, username=email, password=password)
        if not user:
            # Fallback if username wasn't checked by default backend
            try:
                found_user = User.objects.get(email=email)
                if found_user.check_password(password):
                    user = found_user
            except User.DoesNotExist:
                pass

        if not user:
            return Response(
                {"error": "Invalid email or password."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        is_student = StudentProfile.objects.filter(user=user).exists()
        if is_student:
            token, _ = Token.objects.get_or_create(user=user)
            user_data = AuthUserSerializer(user).data
            return Response({
                "token": token.key,
                "user": user_data,
            })

        # Staff or platform admin: require an emailed one-time code unless
        # OTP login is disabled (e.g. local development).
        if settings.OTP_LOGIN_ENABLED:
            otp = issue_login_otp(user)
            if otp is None:
                return Response(
                    {"error": "We couldn't email your secure code. Please try again."},
                    status=status.HTTP_503_SERVICE_UNAVAILABLE,
                )
            return Response({
                "requires_otp": True,
                "email": mask_email(user.email),
                "resend_after": 30,
                "expires_in": OTP_LIFETIME_SECONDS,
            })

        token, _ = Token.objects.get_or_create(user=user)
        user_data = AuthUserSerializer(user).data
        return Response({
            "token": token.key,
            "user": user_data,
        })


class AuthVerifyOtpView(APIView):
    """Exchanges an emailed one-time code for a full auth token."""

    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get("email", "").strip().lower()
        code = request.data.get("code", "").strip()
        if not email or not code:
            return Response(
                {"error": "Email and code are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        otp, error = verify_login_otp(email, code)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)
        user = otp.user
        token, _ = Token.objects.get_or_create(user=user)
        user_data = AuthUserSerializer(user).data
        return Response({
            "token": token.key,
            "user": user_data,
        })


class AuthResendOtpView(APIView):
    """Resends a fresh one-time code after a short cooldown."""

    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get("email", "").strip().lower()
        if not email:
            return Response(
                {"error": "Email is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        otp, error, resend_after = resend_login_otp(email)
        if error:
            return Response(
                {"error": error, "resend_after": resend_after},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response({
            "requires_otp": True,
            "email": mask_email(email),
            "resend_after": resend_after,
            "expires_in": OTP_LIFETIME_SECONDS,
        })


class ForgotPasswordView(APIView):
    """Sends a one-time code to the account email for password reset."""

    permission_classes = [AllowAny]

    def post(self, request):
        serializer = ForgotPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data["email"].lower().strip()

        try:
            user = User.objects.get(email__iexact=email)
        except User.DoesNotExist:
            # Do not reveal whether an account exists for this email.
            return Response({
                "requires_otp": True,
                "email": mask_email(email),
                "resend_after": 30,
                "expires_in": OTP_LIFETIME_SECONDS,
            })

        otp = issue_login_otp(user)
        if otp is None:
            return Response(
                {"error": "We couldn't email your reset code. Please try again."},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        return Response({
            "requires_otp": True,
            "email": mask_email(email),
            "resend_after": 30,
            "expires_in": OTP_LIFETIME_SECONDS,
        })


class ResetPasswordView(APIView):
    """Verifies the emailed one-time code and sets a new password."""

    permission_classes = [AllowAny]

    def post(self, request):
        serializer = ResetPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data["email"].lower().strip()
        code = serializer.validated_data["code"].strip()
        new_password = serializer.validated_data["new_password"]

        otp, error = verify_login_otp(email, code)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)

        user = otp.user
        user.set_password(new_password)
        user.save(update_fields=["password"])
        # Invalidate existing tokens so the new password is enforced.
        Token.objects.filter(user=user).delete()
        return Response({"status": "ok", "message": "Password updated successfully."})


class AuthMeView(APIView):
    """Returns current authenticated user profile."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        staff = getattr(user, "staff_profile", None)
        if staff and staff.institution:
            inst = staff.institution
            if inst.status != InstitutionStatus.ACTIVE and inst.invoices.filter(status=InvoiceStatus.PAID).exists():
                inst.status = InstitutionStatus.ACTIVE
                inst.save(update_fields=["status", "updated_at"])
        user_data = AuthUserSerializer(user).data
        return Response(user_data)


class AuthLogoutView(APIView):
    """Invalidates the auth token on logout."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        Token.objects.filter(user=request.user).delete()
        return Response({"status": "ok", "message": "Successfully logged out."})


class IsPlatformAdmin(permissions.BasePermission):
    """Allows access only to platform super administrators."""

    def has_permission(self, request, view):
        user = request.user
        return bool(user and user.is_authenticated and user.is_superuser)


class PlatformAdminOverviewView(APIView):
    """Platform-wide overview for system super administrators.

    Aggregates every institution tenant, subscription invoice and payment
    status into a single payload powering the platform admin dashboard.
    """

    permission_classes = [IsPlatformAdmin]

    def get(self, request):
        institutions = list(
            Institution.objects.annotate(
                students_count=Count("students", distinct=True),
                staff_count=Count("staff_members", distinct=True),
                divisions_count=Count("divisions", distinct=True),
                departments_count=Count("departments", distinct=True),
                programs_count=Count("programs", distinct=True),
            ).order_by("name")
        )

        invoices = list(
            InstitutionInvoice.objects.select_related("institution", "plan").order_by("-created_at")
        )

        institutions_by_status = {s.value: 0 for s in InstitutionStatus}
        for inst in institutions:
            institutions_by_status[inst.status] = institutions_by_status.get(inst.status, 0) + 1

        invoices_by_status = {s.value: 0 for s in InvoiceStatus}
        plan_counts: dict[str, int] = {}
        for inv in invoices:
            invoices_by_status[inv.status] = invoices_by_status.get(inv.status, 0) + 1
            plan_key = inv.plan_name or "Unassigned"
            plan_counts[plan_key] = plan_counts.get(plan_key, 0) + 1

        paid_total = Decimal("0.00")
        billed_total = Decimal("0.00")
        outstanding_total = Decimal("0.00")
        for inv in invoices:
            billed_total += inv.total_amount
            if inv.status == InvoiceStatus.PAID:
                paid_total += inv.total_amount
            elif inv.status in (InvoiceStatus.UNPAID, InvoiceStatus.PAYMENT_SUBMITTED):
                outstanding_total += inv.total_amount

        latest_by_institution = {}
        for inv in invoices:
            latest_by_institution.setdefault(str(inv.institution_id), inv)

        institution_rows = []
        for inst in institutions:
            latest = latest_by_institution.get(str(inst.id))
            institution_rows.append(
                {
                    "id": str(inst.id),
                    "name": inst.name,
                    "short_name": inst.short_name,
                    "slug": inst.slug,
                    "institution_type": inst.institution_type,
                    "institution_type_display": inst.get_institution_type_display(),
                    "ownership": inst.ownership,
                    "ownership_display": inst.get_ownership_display(),
                    "regulator": inst.regulator,
                    "regulator_display": inst.get_regulator_display(),
                    "state": inst.state,
                    "status": inst.status,
                    "status_display": inst.get_status_display(),
                    "is_founding_partner": inst.is_founding_partner,
                    "created_at": inst.created_at.isoformat(),
                    "students_count": inst.students_count,
                    "staff_count": inst.staff_count,
                    "divisions_count": inst.divisions_count,
                    "departments_count": inst.departments_count,
                    "programs_count": inst.programs_count,
                    "latest_invoice": (
                        {
                            "invoice_number": latest.invoice_number,
                            "total_amount": float(latest.total_amount),
                            "currency": latest.currency,
                            "status": latest.status,
                            "status_display": latest.get_status_display(),
                        }
                        if latest
                        else None
                    ),
                }
            )

        recent_invoices = []
        for inv in invoices[:8]:
            recent_invoices.append(
                {
                    "id": str(inv.id),
                    "invoice_number": inv.invoice_number,
                    "institution": str(inv.institution_id),
                    "institution_name": inv.institution.name,
                    "institution_short_name": inv.institution.short_name,
                    "institution_status": inv.institution.status,
                    "plan_name": inv.plan_name,
                    "total_amount": float(inv.total_amount),
                    "currency": inv.currency,
                    "status": inv.status,
                    "status_display": inv.get_status_display(),
                    "due_date": inv.due_date.isoformat() if inv.due_date else None,
                    "payment_reference": inv.payment_reference or None,
                    "payment_submitted_at": (
                        inv.payment_submitted_at.isoformat() if inv.payment_submitted_at else None
                    ),
                    "confirmed_at": inv.confirmed_at.isoformat() if inv.confirmed_at else None,
                    "created_at": inv.created_at.isoformat(),
                }
            )

        return Response(
            {
                "totals": {
                    "institutions": len(institutions),
                    "users": User.objects.filter(is_active=True).count(),
                    "staff": InstitutionStaff.objects.filter(is_active=True).count(),
                    "students": StudentProfile.objects.count(),
                    "divisions": AcademicDivision.objects.count(),
                    "departments": Department.objects.count(),
                    "programs": AcademicProgram.objects.count(),
                    "pathways": Pathway.objects.count(),
                    "invoices": len(invoices),
                },
                "institutions_by_status": institutions_by_status,
                "invoices_by_status": invoices_by_status,
                "plans": plan_counts,
                "revenue": {
                    "total_billed": float(billed_total),
                    "total_paid": float(paid_total),
                    "outstanding": float(outstanding_total),
                    "currency": "NGN",
                },
                "institutions": institution_rows,
                "recent_invoices": recent_invoices,
            }
        )


class AdminBankDetailViewSet(viewsets.ModelViewSet):
    """Admin console: full CRUD over company bank accounts for invoices."""

    queryset = CompanyBankDetail.objects.all().order_by("-is_active", "-created_at")
    serializer_class = CompanyBankDetailSerializer
    permission_classes = [IsPlatformAdmin]


class AdminPricingPlanViewSet(viewsets.ModelViewSet):
    """Admin console: full CRUD over pricing plans & fee structures."""

    queryset = PricingPlan.objects.all().order_by("base_fee")
    serializer_class = PricingPlanSerializer
    permission_classes = [IsPlatformAdmin]


class AdminInvoiceViewSet(viewsets.ReadOnlyModelViewSet):
    """Admin console: read-only invoices plus confirm/reject verification actions."""

    queryset = (
        InstitutionInvoice.objects.all()
        .select_related("institution", "plan", "confirmed_by")
        .order_by("-created_at")
    )
    serializer_class = InstitutionInvoiceSerializer
    permission_classes = [IsPlatformAdmin]

    @action(detail=True, methods=["get"], url_path="pdf")
    def download_pdf(self, request, pk=None):
        """Return a professional printable PDF of the invoice (with payment evidence)."""
        invoice = self.get_object()
        pdf_bytes = build_invoice_pdf(invoice)
        filename = f"{invoice.invoice_number}.pdf".replace(" ", "_")
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["post"], url_path="confirm")
    def confirm(self, request, pk=None):
        invoice = self.get_object()
        invoice.status = InvoiceStatus.PAID
        invoice.confirmed_by = request.user
        invoice.confirmed_at = timezone.now()
        invoice.admin_notes = request.data.get("admin_notes", "")
        invoice.save()
        institution = invoice.institution
        if institution.status != InstitutionStatus.ACTIVE:
            institution.status = InstitutionStatus.ACTIVE
            institution.save(update_fields=["status", "updated_at"])
        return Response(
            InstitutionInvoiceSerializer(invoice, context={"request": request}).data
        )

    @action(detail=True, methods=["post"], url_path="reject")
    def reject(self, request, pk=None):
        invoice = self.get_object()
        invoice.status = InvoiceStatus.REJECTED
        invoice.confirmed_by = request.user
        invoice.confirmed_at = timezone.now()
        invoice.admin_notes = request.data.get("admin_notes", "")
        invoice.save()
        institution = invoice.institution
        if institution.status != InstitutionStatus.REJECTED:
            institution.status = InstitutionStatus.REJECTED
            institution.save(update_fields=["status", "updated_at"])
        return Response(
            InstitutionInvoiceSerializer(invoice, context={"request": request}).data
        )


class AdminUserViewSet(viewsets.ReadOnlyModelViewSet):
    """Admin console: browse every platform user with their roles."""

    serializer_class = AdminUserSerializer
    permission_classes = [IsPlatformAdmin]

    def get_queryset(self):
        qs = User.objects.all().order_by("-date_joined")
        search = self.request.query_params.get("search")
        if search:
            qs = qs.filter(Q(email__icontains=search) | Q(name__icontains=search))
        return qs.prefetch_related(
            "institution_staff_profiles", "student_profile", "staff_assignments"
        )


class AdminInstitutionDetailView(APIView):
    """Platform admin drill-down into a single institution tenant.

    Returns the tenant profile plus every related record (students, staff,
    divisions/departments/programs, pathways, invoices) in one prefetched
    payload so the admin console renders fast without N+1 queries.
    """

    permission_classes = [IsPlatformAdmin]

    def get(self, request, institution_id):
        inst = (
            Institution.objects.prefetch_related(
                "students",
                "students__program__department__division",
                "students__entry_session",
                "staff_members",
                "staff_members__user",
                "staff_members__division",
                "staff_members__department",
                "divisions",
                "divisions__departments",
                "divisions__departments__programs",
                "programs",
                "programs__department__division",
                "pathways",
                "pathways__program__department__division",
                "pathways__created_by",
                "invoices",
                "invoices__plan",
                "invoices__confirmed_by",
            )
            .filter(id=institution_id)
            .first()
        )
        if inst is None:
            return Response(
                {"detail": "Institution not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        students = inst.students.all()
        staff = inst.staff_members.all()
        divisions = inst.divisions.all()
        programs = inst.programs.all()
        pathways = inst.pathways.all()
        invoices = inst.invoices.all()
        departments = Department.objects.filter(institution=inst)

        return Response(
            {
                "id": str(inst.id),
                "name": inst.name,
                "short_name": inst.short_name,
                "slug": inst.slug,
                "institution_type": inst.institution_type,
                "institution_type_display": inst.get_institution_type_display(),
                "ownership": inst.ownership,
                "ownership_display": inst.get_ownership_display(),
                "regulator": inst.regulator,
                "regulator_display": inst.get_regulator_display(),
                "tier_two_term": inst.tier_two_term,
                "state": inst.state,
                "address": inst.address,
                "domain_whitelist": inst.domain_whitelist,
                "is_founding_partner": inst.is_founding_partner,
                "status": inst.status,
                "status_display": inst.get_status_display(),
                "created_at": inst.created_at.isoformat(),
                "updated_at": inst.updated_at.isoformat(),
                "totals": {
                    "students": students.count(),
                    "staff": staff.count(),
                    "divisions": divisions.count(),
                    "departments": departments.count(),
                    "programs": programs.count(),
                    "pathways": pathways.count(),
                    "invoices": invoices.count(),
                },
                "students": StudentProfileSerializer(
                    students, many=True, context={"request": request}
                ).data,
                "staff": InstitutionStaffSerializer(
                    staff, many=True, context={"request": request}
                ).data,
                "divisions": AcademicDivisionSerializer(
                    divisions, many=True, context={"request": request}
                ).data,
                "programs": AcademicProgramSerializer(
                    programs, many=True, context={"request": request}
                ).data,
                "pathways": PathwayListSerializer(
                    pathways, many=True, context={"request": request}
                ).data,
                "invoices": InstitutionInvoiceSerializer(
                    invoices, many=True, context={"request": request}
                ).data,
            }
        )


class AdminInstitutionStatusView(APIView):
    """Platform admin: deactivate (suspend) or reactivate an institution tenant."""

    permission_classes = [IsPlatformAdmin]

    def post(self, request, institution_id, action):
        if action not in ("deactivate", "reactivate"):
            return Response(
                {"detail": "Invalid action."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        inst = Institution.objects.filter(id=institution_id).first()
        if inst is None:
            return Response(
                {"detail": "Institution not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if action == "deactivate":
            if inst.status == InstitutionStatus.SUSPENDED:
                return Response(
                    {"detail": "Institution is already deactivated.", "status": inst.status},
                    status=status.HTTP_409_CONFLICT,
                )
            inst.status = InstitutionStatus.SUSPENDED
            detail = "Institution deactivated."
        else:
            inst.status = InstitutionStatus.ACTIVE
            detail = "Institution reactivated."

        inst.save(update_fields=["status", "updated_at"])
        return Response(
            {
                "detail": detail,
                "status": inst.status,
                "status_display": inst.get_status_display(),
            }
        )


def get_staff_scoped_students(user, institution_id=None):
    """
    Returns StudentProfile QuerySet restricted strictly to the staff member's
    assigned department(s) or faculty/division(s).
    """
    if not user.is_authenticated:
        qs = StudentProfile.objects.all()
        if institution_id:
            qs = qs.filter(institution_id=institution_id)
        return qs.select_related(
            "user", "program", "program__department", "program__department__division", "institution", "entry_session"
        )

    # Check staff assignments
    assignments = user.staff_assignments.filter(is_active=True)

    # Check fallback legacy InstitutionStaff if no granular assignments
    if not assignments.exists():
        legacy_staff = user.institution_staff_profiles.filter(is_active=True).first()
        if legacy_staff:
            if legacy_staff.role in ["SUPERADMIN", "DIRECTOR_CAREER_SERVICES"]:
                qs = StudentProfile.objects.filter(institution=legacy_staff.institution)
            elif legacy_staff.department:
                qs = StudentProfile.objects.filter(program__department=legacy_staff.department)
            elif legacy_staff.division:
                qs = StudentProfile.objects.filter(program__department__division=legacy_staff.division)
            else:
                qs = StudentProfile.objects.filter(institution=legacy_staff.institution)
            if institution_id:
                qs = qs.filter(institution_id=institution_id)
            return qs.select_related(
                "user", "program", "program__department", "program__department__division", "institution", "entry_session"
            )
        # If user is a student, they can only view themselves
        if hasattr(user, "student_profile") and user.student_profile:
            return StudentProfile.objects.filter(id=user.student_profile.id)

        # Allow institution filter for public/admin demo queries
        qs = StudentProfile.objects.all()
        if institution_id:
            qs = qs.filter(institution_id=institution_id)
        return qs.select_related(
            "user", "program", "program__department", "program__department__division", "institution", "entry_session"
        )

    # If user is Superadmin or Director of Career Services -> Full Institution Access
    if assignments.filter(role_at_unit__in=[
        StaffRoleAtUnit.SUPERADMIN,
        StaffRoleAtUnit.DIRECTOR_CAREER_SERVICES,
    ]).exists():
        inst_id = institution_id or assignments.first().institution_id
        return StudentProfile.objects.filter(institution_id=inst_id).select_related(
            "user", "program", "program__department", "program__department__division", "institution", "entry_session"
        )

    # Compile scoped divisions and departments
    scoped_division_ids = assignments.filter(department__isnull=True, division__isnull=False).values_list("division_id", flat=True)
    scoped_department_ids = assignments.filter(department__isnull=False).values_list("department_id", flat=True)

    q_filter = Q(program__department_id__in=scoped_department_ids) | Q(program__department__division_id__in=scoped_division_ids)
    qs = StudentProfile.objects.filter(q_filter)
    if institution_id:
        qs = qs.filter(institution_id=institution_id)

    return qs.select_related(
        "user", "program", "program__department", "program__department__division", "institution", "entry_session"
    )


class StaffAssignmentViewSet(viewsets.ModelViewSet):
    """Endpoints for managing fine-grained departmental/division staff assignments."""

    queryset = StaffAssignment.objects.all().select_related("user", "institution", "division", "department")
    serializer_class = StaffAssignmentSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution")
        div_id = self.request.query_params.get("division")
        dept_id = self.request.query_params.get("department")
        role = self.request.query_params.get("role_at_unit")
        user_id = self.request.query_params.get("user")

        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        if div_id:
            qs = qs.filter(division_id=div_id)
        if dept_id:
            qs = qs.filter(department_id=dept_id)
        if role:
            qs = qs.filter(role_at_unit=role.upper())
        if user_id:
            qs = qs.filter(user_id=user_id)
        return qs

    @action(detail=False, methods=["get"], url_path="my-caseload")
    def my_caseload(self, request):
        """Returns currently authenticated staff member's assigned units and student metrics."""
        if not request.user.is_authenticated:
            return Response({"error": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)

        assignments = request.user.staff_assignments.filter(is_active=True).select_related("institution", "division", "department")
        students_qs = get_staff_scoped_students(request.user)

        total_students = students_qs.count()
        siwes_qualifying = students_qs.filter(siwes_clearance_status="QUALIFYING").count()
        final_year_count = sum(1 for s in students_qs if s.is_final_year)

        return Response({
            "assignments": StaffAssignmentSerializer(assignments, many=True).data,
            "metrics": {
                "total_managed_students": total_students,
                "siwes_qualifying_candidates": siwes_qualifying,
                "final_year_students": final_year_count,
            },
        })


class StudentProfileViewSet(viewsets.ModelViewSet):
    """
    Endpoints for student identities anchored to Tier-4 AcademicProgram.
    Scoped strictly to staff member's assigned departments.
    """

    queryset = StudentProfile.objects.all().select_related(
        "user", "program", "program__department", "program__department__division", "institution", "entry_session"
    )
    serializer_class = StudentProfileSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        inst_id = self.request.query_params.get("institution")
        qs = get_staff_scoped_students(self.request.user, inst_id)

        prog_id = self.request.query_params.get("program")
        dept_id = self.request.query_params.get("department")
        div_id = self.request.query_params.get("division")
        year = self.request.query_params.get("year_of_study")
        siwes_status = self.request.query_params.get("siwes_status")
        standing = self.request.query_params.get("academic_standing")
        search = self.request.query_params.get("search")

        if prog_id:
            qs = qs.filter(program_id=prog_id)
        if dept_id:
            qs = qs.filter(program__department_id=dept_id)
        if div_id:
            qs = qs.filter(program__department__division_id=div_id)
        if year:
            qs = qs.filter(year_of_study=year)
        if siwes_status:
            qs = qs.filter(siwes_clearance_status=siwes_status.upper())
        if standing:
            qs = qs.filter(academic_standing=standing.upper())
        if search:
            qs = qs.filter(
                Q(matric_number__icontains=search)
                | Q(user__name__icontains=search)
                | Q(user__email__icontains=search)
                | Q(program__name__icontains=search)
            )
        return qs

    def create(self, request, *args, **kwargs):
        """Creates a student user account and initializes StudentProfile."""
        serializer = StudentProfileCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        email = data["email"].lower().strip()
        matric = data["matric_number"].strip()

        # Matric number is unique per institution — check before insert so the
        # caller gets a clear message instead of a DB IntegrityError (500).
        existing_user = User.objects.filter(email=email).first()
        if StudentProfile.objects.filter(
            institution_id=data["institution"],
            matric_number__iexact=matric,
        ).exclude(user_id=existing_user.id if existing_user else -1).exists():
            return duplicate_error(
                f"A student with matric number '{matric}' is already registered in this institution. "
                "Matric numbers must be unique per institution."
            )

        with transaction.atomic():
            user, created = User.objects.get_or_create(
                email=email,
                defaults={"name": data["name"]},
            )
            if created or not user.has_usable_password():
                user.set_password(data.get("password", "1234!@#$"))
                user.name = data["name"]
                user.save()

            program = AcademicProgram.objects.get(id=data["program"])
            institution = Institution.objects.get(id=data["institution"])
            session = AcademicSession.objects.get(id=data["entry_session"])

            student, _ = StudentProfile.objects.update_or_create(
                user=user,
                defaults={
                    "institution": institution,
                    "program": program,
                    "matric_number": matric,
                    "jamb_reg_number": data.get("jamb_reg_number", "").strip(),
                    "entry_session": session,
                    "entry_mode": data.get("entry_mode", "UTME"),
                    "year_of_study": data.get("year_of_study", 1),
                    "cgpa": data.get("cgpa"),
                    "phone_number": data.get("phone_number", ""),
                    "state_of_origin": data.get("state_of_origin", ""),
                    "gender": data.get("gender", ""),
                    "portfolio_url": data.get("portfolio_url", ""),
                    "is_verified_student": True,
                },
            )

        output_serializer = StudentProfileSerializer(student)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["get"], url_path="me")
    def me(self, request):
        """Returns the currently logged-in student's complete profile."""
        if not request.user.is_authenticated:
            return Response({"error": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)
        if not hasattr(request.user, "student_profile") or not request.user.student_profile:
            return Response({"error": "User does not have a student profile"}, status=status.HTTP_404_NOT_FOUND)

        serializer = StudentProfileSerializer(request.user.student_profile)
        return Response(serializer.data)

        student.save()
        return Response(StudentProfileSerializer(student).data)

    @action(detail=True, methods=["post"], url_path="generate-credentials")
    def generate_credentials(self, request, id=None):
        """Generates a secure password and emails login credentials to the student via Mailpit SMTP."""
        student = self.get_object()
        serializer = StudentCredentialGenerationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        result = StudentCredentialService.generate_and_dispatch_credentials(
            student_profile_id=str(student.id),
            custom_password=data.get("custom_password"),
            login_url=data.get("login_url", "http://localhost:5173"),
        )
        return Response(result, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="enroll-pathway")
    def enroll_pathway(self, request, id=None):
        """Enrolls the student in an active career pathway for their degree program."""
        student = self.get_object()
        serializer = StudentEnrollPathwaySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        pathway_id = serializer.validated_data["pathway"]

        try:
            pathway = Pathway.objects.get(id=pathway_id)
        except Pathway.DoesNotExist:
            return Response({"error": "Pathway not found"}, status=status.HTTP_404_NOT_FOUND)

        student.active_pathway = pathway
        student.save(update_fields=["active_pathway", "updated_at"])
        student.recalculate_employability()

        return Response(StudentProfileSerializer(student).data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="me/dashboard")
    def my_dashboard(self, request):
        """Returns the logged-in student's complete dashboard data: profile, enrolled pathway roadmap, submissions, and employability score."""
        if not request.user.is_authenticated:
            return Response({"error": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)
        if not hasattr(request.user, "student_profile") or not request.user.student_profile:
            return Response({"error": "User does not have a student profile"}, status=status.HTTP_404_NOT_FOUND)

        student = (
            StudentProfile.objects.select_related(
                "user",
                "institution",
                "program",
                "program__department",
                "program__department__division",
                "active_pathway",
            )
            .get(id=request.user.student_profile.id)
        )

        employability_summary = student.recalculate_employability()

        active_pathway_data = None
        if student.active_pathway:
            active_pathway_data = PathwayDetailSerializer(student.active_pathway).data
        else:
            # Fallback: find the first active pathway for the student's program
            default_pathway = (
                Pathway.objects.filter(program=student.program, is_active=True)
                .prefetch_related("milestones")
                .first()
            )
            if default_pathway:
                student.active_pathway = default_pathway
                student.save(update_fields=["active_pathway"])
                employability_summary = student.recalculate_employability()
                active_pathway_data = PathwayDetailSerializer(default_pathway).data

        submissions_qs = (
            StudentMilestoneSubmission.objects.filter(student=student)
            .select_related("milestone", "reviewed_by")
            .order_by("-created_at")
        )
        submissions_data = StudentMilestoneSubmissionSerializer(submissions_qs, many=True).data

        return Response({
            "profile": StudentProfileSerializer(student).data,
            "active_pathway": active_pathway_data,
            "submissions": submissions_data,
            "employability_summary": employability_summary,
        })

    @action(detail=True, methods=["get"], url_path="dossier")
    def student_dossier(self, request, id=None):
        """Assembles and returns a complete 360° student dossier for departmental counsellors and HODs."""
        student = self.get_object()
        submissions = student.milestone_submissions.select_related("milestone").order_by("milestone__order_index")
        assessments = student.assessment_sessions.select_related("assessment").order_by("-completed_at")
        counselling_sessions = student.counselling_sessions.select_related("counsellor", "counsellor__user").prefetch_related("case_notes").order_by("-created_at")
        case_notes = student.counsellor_case_notes.select_related("author", "author__user").order_by("-created_at")

        latest_ai_conv = student.ai_conversations.order_by("-updated_at").first()
        ai_summary = latest_ai_conv.case_summary if latest_ai_conv else ""

        emp_summary = {
            "employability_score": float(student.employability_score),
            "verified_points_total": student.verified_points_total,
            "milestones_completed_count": student.milestones_completed_count,
            "cgpa": float(student.cgpa) if student.cgpa else 0.0,
            "is_siwes_year": student.is_siwes_year,
            "academic_standing": student.get_academic_standing_display(),
        }

        dossier_data = {
            "profile": StudentProfileSerializer(student).data,
            "active_pathway": PathwayDetailSerializer(student.active_pathway).data if student.active_pathway else None,
            "submissions": StudentMilestoneSubmissionSerializer(submissions, many=True).data,
            "assessments": StudentAssessmentSessionSerializer(assessments, many=True).data,
            "counselling_sessions": CounsellingSessionSerializer(counselling_sessions, many=True).data,
            "case_notes": CounsellingCaseNoteSerializer(case_notes, many=True).data,
            "ai_coach_summary": ai_summary,
            "employability_summary": emp_summary,
        }
        return Response(dossier_data, status=status.HTTP_200_OK)




class PathwayViewSet(viewsets.ModelViewSet):
    """CRUD and blueprint template management endpoints for Career Pathways."""

    queryset = (
        Pathway.objects.all()
        .select_related(
            "institution",
            "program",
            "program__department",
            "program__department__division",
            "created_by",
            "cloned_from",
        )
        .prefetch_related("milestones")
    )
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_serializer_class(self):
        if self.action == "retrieve":
            return PathwayDetailSerializer
        if self.action in ["create", "update", "partial_update"]:
            return PathwayCreateSerializer
        return PathwayListSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution")
        prog_id = self.request.query_params.get("program")
        dept_id = self.request.query_params.get("department")
        div_id = self.request.query_params.get("division")
        is_template = self.request.query_params.get("is_template")
        search = self.request.query_params.get("search")

        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        if prog_id:
            qs = qs.filter(program_id=prog_id)
        if dept_id:
            qs = qs.filter(program__department_id=dept_id)
        if div_id:
            qs = qs.filter(program__department__division_id=div_id)
        if is_template is not None:
            qs = qs.filter(is_template=is_template.lower() in ["true", "1"])
        if search:
            qs = qs.filter(Q(title__icontains=search) | Q(career_role__icontains=search) | Q(industry_sector__icontains=search))

        return qs

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        pathway = serializer.save(created_by=user)
        pathway.recalculate_totals()

    @action(detail=True, methods=["post"], url_path="clone")
    def clone_pathway(self, request, id=None):
        """1-Click cloning: Clones a master template into a customized pathway for an academic program."""
        serializer = PathwayClonePayloadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        try:
            target_program = AcademicProgram.objects.get(id=data["target_program"])
        except AcademicProgram.DoesNotExist:
            return Response({"error": "Target program not found"}, status=status.HTTP_404_NOT_FOUND)

        user = request.user if request.user.is_authenticated else None

        new_pathway = PathwayTemplateService.clone_template_to_program(
            template_id=id,
            target_program=target_program,
            user=user,
            custom_title=data.get("custom_title"),
            custom_description=data.get("custom_description"),
        )

        return Response(
            PathwayDetailSerializer(new_pathway).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="publish-as-template")
    def publish_as_template(self, request, id=None):
        """Publishes an active custom pathway as a master blueprint template."""
        serializer = PathwayPublishTemplatePayloadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        visibility = serializer.validated_data.get("visibility", "INSTITUTION")

        pathway = PathwayTemplateService.publish_as_template(
            pathway_id=id,
            visibility=visibility,
        )
        return Response(PathwayDetailSerializer(pathway).data)

    @action(detail=False, methods=["get"], url_path="templates")
    def templates_catalog(self, request):
        """Returns all reusable blueprint templates, optionally filtered by award level or sector."""
        qs = Pathway.objects.filter(is_template=True).select_related(
            "institution",
            "program",
            "program__department",
            "created_by",
        ).prefetch_related("milestones")

        award_level = request.query_params.get("award_level")
        if award_level:
            qs = qs.filter(program__award_level=award_level)

        serializer = PathwayListSerializer(qs, many=True)
        return Response(serializer.data)


class PathwayMilestoneViewSet(viewsets.ModelViewSet):
    """CRUD and reordering endpoints for individual Pathway Milestones."""

    queryset = PathwayMilestone.objects.all().select_related("pathway", "pathway__program")
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return PathwayMilestoneCreateUpdateSerializer
        return PathwayMilestoneSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        pathway_id = self.request.query_params.get("pathway")
        if pathway_id:
            qs = qs.filter(pathway_id=pathway_id)
        return qs

    def perform_create(self, serializer):
        milestone = serializer.save()
        milestone.pathway.recalculate_totals()

    def perform_update(self, serializer):
        milestone = serializer.save()
        milestone.pathway.recalculate_totals()

    def perform_destroy(self, instance):
        pathway = instance.pathway
        instance.delete()
        pathway.recalculate_totals()

    @action(detail=False, methods=["post"], url_path="reorder")
    def reorder_milestones(self, request):
        """Reorders milestones in a pathway: accepts array of [{id: UUID, order_index: int}]."""
        items = request.data.get("items", [])
        if not items:
            return Response({"error": "No items provided"}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            for item in items:
                PathwayMilestone.objects.filter(id=item["id"]).update(order_index=item["order_index"])

        return Response({"status": "ok", "message": "Milestones successfully reordered."})


class StudentMilestoneSubmissionViewSet(viewsets.ModelViewSet):
    """CRUD and evaluation review endpoints for Student Milestone Submissions."""

    queryset = (
        StudentMilestoneSubmission.objects.all()
        .select_related(
            "student",
            "student__user",
            "student__program",
            "student__institution",
            "milestone",
            "milestone__pathway",
            "reviewed_by",
        )
    )
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_serializer_class(self):
        if self.action == "create":
            return StudentSubmissionCreateSerializer
        return StudentMilestoneSubmissionSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        student_id = self.request.query_params.get("student")
        milestone_id = self.request.query_params.get("milestone")
        pathway_id = self.request.query_params.get("pathway")
        inst_id = self.request.query_params.get("institution")
        status_filter = self.request.query_params.get("status")

        if student_id:
            qs = qs.filter(student_id=student_id)
        if milestone_id:
            qs = qs.filter(milestone_id=milestone_id)
        if pathway_id:
            qs = qs.filter(milestone__pathway_id=pathway_id)
        if inst_id:
            qs = qs.filter(student__institution_id=inst_id)
        if status_filter:
            qs = qs.filter(status=status_filter.upper())

        return qs

    def perform_create(self, serializer):
        # Resolve student
        student = None
        if self.request.user.is_authenticated and hasattr(self.request.user, "student_profile") and self.request.user.student_profile:
            student = self.request.user.student_profile
        else:
            student_id = self.request.data.get("student")
            if student_id:
                student = StudentProfile.objects.get(id=student_id)

        if not student:
            raise ValueError("Student profile could not be determined for this submission.")

        milestone = serializer.validated_data["milestone"]

        # Check if already submitted (create or update)
        existing = StudentMilestoneSubmission.objects.filter(student=student, milestone=milestone).first()
        if existing:
            existing.evidence_url = serializer.validated_data.get("evidence_url", existing.evidence_url)
            existing.submission_notes = serializer.validated_data.get("submission_notes", existing.submission_notes)
            existing.status = SubmissionStatus.PENDING_REVIEW
            existing.save()
            student.recalculate_employability()
            return

        submission = serializer.save(student=student, status=SubmissionStatus.PENDING_REVIEW)
        student.recalculate_employability()

    @action(detail=True, methods=["post"], url_path="review")
    def review_submission(self, request, id=None):
        """Staff Review action: Approve (VERIFIED), request changes, or reject + award points."""
        submission = self.get_object()
        serializer = StudentSubmissionReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        submission.status = data["status"]
        submission.review_feedback = data.get("review_feedback", "")
        submission.reviewed_by = request.user if request.user.is_authenticated else None
        submission.reviewed_at = timezone.now()

        if submission.status == SubmissionStatus.VERIFIED:
            # Award points specified or default to full milestone points
            points = data.get("points_awarded")
            submission.points_awarded = points if points is not None else submission.milestone.points
        else:
            submission.points_awarded = 0

        submission.save()

        # Recalculate student's total employability score
        submission.student.recalculate_employability()

        return Response(
            StudentMilestoneSubmissionSerializer(submission).data,
            status=status.HTTP_200_OK,
        )


# =============================================================================
# Diagnostic Assessments ViewSets
# =============================================================================

class DiagnosticAssessmentViewSet(viewsets.ReadOnlyModelViewSet):
    """Catalog of scientific psychometric models, Holland RIASEC tests, and skill diagnostics."""

    queryset = (
        DiagnosticAssessment.objects.filter(is_active=True)
        .prefetch_related("questions")
        .order_by("assessment_type", "title")
    )
    permission_classes = [AllowAny]
    lookup_field = "slug"

    def get_serializer_class(self):
        if self.action == "retrieve":
            return DiagnosticAssessmentDetailSerializer
        return DiagnosticAssessmentListSerializer

    @action(detail=True, methods=["get"], url_path="questions")
    def questions(self, request, slug=None):
        """Returns all questions and choices for this assessment."""
        assessment = self.get_object()
        questions = assessment.questions.all().order_by("order_index")
        serializer = DiagnosticQuestionSerializer(questions, many=True)
        return Response(serializer.data)


class StudentAssessmentSessionViewSet(viewsets.ModelViewSet):
    """Execution and historical results for student diagnostic tests."""

    queryset = (
        StudentAssessmentSession.objects.all()
        .select_related("student", "student__user", "assessment")
        .order_by("-completed_at", "-started_at")
    )
    serializer_class = StudentAssessmentSessionSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def create(self, request, *args, **kwargs):
        """Submits assessment answers, executes psychometric algorithm, and saves completed session."""
        serializer = StudentAssessmentSubmitSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        assessment_id = data["assessment_id"]
        raw_responses = data["raw_responses"]

        # Identify student
        student = None
        if request.user.is_authenticated and hasattr(request.user, "student_profile") and request.user.student_profile:
            student = request.user.student_profile
        else:
            student_id = request.data.get("student_id")
            if student_id:
                student = StudentProfile.objects.get(id=student_id)
            else:
                student = StudentProfile.objects.first()

        if not student:
            return Response({"error": "Student profile not found"}, status=status.HTTP_400_BAD_REQUEST)

        session = PsychometricService.evaluate_and_save_session(
            student_id=str(student.id),
            assessment_id=str(assessment_id),
            raw_responses=raw_responses,
        )

        return Response(
            StudentAssessmentSessionSerializer(session).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=["get"], url_path="my-results")
    def my_results(self, request):
        """Returns all completed diagnostic assessment results for the current student."""
        student = None
        if request.user.is_authenticated and hasattr(request.user, "student_profile") and request.user.student_profile:
            student = request.user.student_profile
        else:
            student = StudentProfile.objects.first()

        if not student:
            return Response([])

        sessions = (
            StudentAssessmentSession.objects.filter(student=student, status="COMPLETED")
            .select_related("assessment")
            .order_by("-completed_at")
        )
        return Response(StudentAssessmentSessionSerializer(sessions, many=True).data)


# =============================================================================
# 24/7 AI Career Coach ViewSet
# =============================================================================

class AICoachViewSet(viewsets.ViewSet):
    """24/7 AI Career Coaching endpoints grounded in institutional handbooks and student dossier."""

    permission_classes = [AllowAny]

    def _get_student(self, request):
        if request.user.is_authenticated and hasattr(request.user, "student_profile") and request.user.student_profile:
            return request.user.student_profile
        student_id = request.query_params.get("student_id") or request.data.get("student_id")
        if student_id:
            try:
                return StudentProfile.objects.get(id=student_id)
            except StudentProfile.DoesNotExist:
                pass
        return StudentProfile.objects.first()

    @action(detail=False, methods=["get", "post"], url_path="conversations")
    def conversations(self, request):
        """List or create AI Coach conversation threads for the current student."""
        student = self._get_student(request)
        if not student:
            return Response({"error": "Student profile not found"}, status=status.HTTP_400_BAD_REQUEST)

        if request.method == "POST":
            title = request.data.get("title", "Career & SIWES Advisory Session")
            conv = AICoachConversation.objects.create(
                student=student,
                title=title,
                is_active=True,
            )
            # Create initial welcoming message from assistant
            initial_content = (
                f"Hello {student.user.name or 'Student'}! I am your 24/7 AI Career Coach at **{student.institution.name}**. "
                f"I am grounded in your official student handbook, departmental SIWES guidelines, and your active "
                f"**{student.active_pathway.title if student.active_pathway else 'Career Pathway'}** roadmap.\n\n"
                f"How can I assist you with your cover letters, milestone evidence, or SIWES preparations today?"
            )
            AICoachMessage.objects.create(
                conversation=conv,
                role="assistant",
                content=initial_content,
            )
            return Response(AICoachConversationSerializer(conv).data, status=status.HTTP_201_CREATED)

        conversations = (
            AICoachConversation.objects.filter(student=student)
            .prefetch_related("messages")
            .order_by("-updated_at")
        )
        return Response(AICoachConversationSerializer(conversations, many=True).data)

    @action(detail=True, methods=["get", "post"], url_path="messages")
    def messages(self, request, pk=None):
        """Get messages for a conversation or send a new student question to the grounded AI Coach."""
        try:
            conversation = AICoachConversation.objects.get(id=pk)
        except AICoachConversation.DoesNotExist:
            return Response({"error": "Conversation not found"}, status=status.HTTP_404_NOT_FOUND)

        if request.method == "POST":
            serializer = AICoachChatPayloadSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            user_msg = serializer.validated_data["message"]

            response_data = StudentAICoachService.ask_coach(
                conversation_id=str(conversation.id),
                user_message=user_msg,
            )
            return Response(response_data, status=status.HTTP_200_OK)

        messages = conversation.messages.all().order_by("created_at")
        return Response(AICoachMessageSerializer(messages, many=True).data)


# =============================================================================
# Seamless Counsellor Handoff & Booking ViewSets
# =============================================================================

class CounsellingSessionViewSet(viewsets.ModelViewSet):
    """Appointment booking and session lifecycle management for students and counsellors."""

    queryset = (
        CounsellingSession.objects.all()
        .select_related(
            "student",
            "student__user",
            "student__program",
            "counsellor",
            "counsellor__user",
        )
        .prefetch_related("case_notes")
        .order_by("-preferred_date", "-created_at")
    )
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_serializer_class(self):
        if self.action == "create":
            return CounsellingSessionCreateSerializer
        return CounsellingSessionSerializer

    def perform_create(self, serializer):
        student = None
        if self.request.user.is_authenticated and hasattr(self.request.user, "student_profile") and self.request.user.student_profile:
            student = self.request.user.student_profile
        else:
            student_id = self.request.data.get("student_id")
            if student_id:
                student = StudentProfile.objects.get(id=student_id)
            else:
                student = StudentProfile.objects.first()

        if not student:
            raise ValueError("Student profile could not be determined for this counselling appointment.")

        serializer.save(student=student, status="REQUESTED")

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(
            CounsellingSessionSerializer(serializer.instance).data,
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    @action(detail=False, methods=["get"], url_path="my-sessions")
    def my_sessions(self, request):
        """Returns sessions for the logged-in student or assigned counsellor."""
        if request.user.is_authenticated and hasattr(request.user, "student_profile") and request.user.student_profile:
            sessions = self.get_queryset().filter(student=request.user.student_profile)
        elif request.user.is_authenticated and hasattr(request.user, "staff_profile") and request.user.staff_profile:
            sessions = self.get_queryset().filter(counsellor=request.user.staff_profile)
        else:
            # Fallback for demo
            student = StudentProfile.objects.first()
            sessions = self.get_queryset().filter(student=student) if student else self.get_queryset()

        return Response(CounsellingSessionSerializer(sessions, many=True).data)

    @action(detail=False, methods=["get"], url_path="available-counsellors")
    def available_counsellors(self, request):
        """Lists available staff counsellors and HODs for appointment booking."""
        institution_id = request.query_params.get("institution")
        department_id = request.query_params.get("department")

        qs = InstitutionStaff.objects.filter(is_active=True).select_related("user", "institution")
        if institution_id:
            qs = qs.filter(institution_id=institution_id)
        if department_id:
            qs = qs.filter(assignments__department_id=department_id)

        counsellors_data = []
        for staff in qs.distinct()[:10]:
            counsellors_data.append({
                "id": str(staff.id),
                "name": staff.user.name,
                "email": staff.user.email,
                "title": staff.title,
                "phone": getattr(staff, "phone_number", ""),
                "office_location": getattr(staff, "office_location", "Department Career Advisory Office"),
                "institution": staff.institution.name,
            })
        return Response(counsellors_data)

    @action(detail=True, methods=["post"], url_path="confirm")
    def confirm_session(self, request, id=None):
        """Counsellor confirms session and designates meeting venue or virtual call link."""
        session = self.get_object()
        serializer = CounsellingSessionConfirmSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        session.status = data.get("status", "CONFIRMED")
        if data.get("scheduled_datetime"):
            session.scheduled_datetime = data["scheduled_datetime"]
        if data.get("meeting_location"):
            session.meeting_location = data["meeting_location"]

        if request.user.is_authenticated and hasattr(request.user, "staff_profile") and request.user.staff_profile:
            session.counsellor = request.user.staff_profile

        session.save()
        return Response(CounsellingSessionSerializer(session).data, status=status.HTTP_200_OK)


class CounsellingCaseNoteViewSet(viewsets.ModelViewSet):
    """Confidential case notes documented by counsellors and linked to student profiles."""

    queryset = (
        CounsellingCaseNote.objects.all()
        .select_related("student", "author", "author__user", "session")
        .order_by("-created_at")
    )
    serializer_class = CounsellingCaseNoteSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def perform_create(self, serializer):
        author = None
        if self.request.user.is_authenticated and hasattr(self.request.user, "staff_profile") and self.request.user.staff_profile:
            author = self.request.user.staff_profile
        else:
            author_id = self.request.data.get("author_id")
            if author_id:
                author = InstitutionStaff.objects.get(id=author_id)
            else:
                author = InstitutionStaff.objects.first()

        student_id = self.request.data.get("student")
        student = StudentProfile.objects.get(id=student_id)

        session_id = self.request.data.get("session")
        session = CounsellingSession.objects.get(id=session_id) if session_id else None

        serializer.save(
            student=student,
            author=author,
            session=session,
        )


class CompanyBankDetailViewSet(viewsets.ReadOnlyModelViewSet):
    """Provides active company bank details for invoice payments."""

    queryset = CompanyBankDetail.objects.filter(is_active=True)
    serializer_class = CompanyBankDetailSerializer
    permission_classes = [AllowAny]

    @action(detail=False, methods=["get"], url_path="active")
    def active(self, request):
        bank = CompanyBankDetail.objects.filter(is_active=True).first()
        if not bank:
            bank = CompanyBankDetail.objects.create(
                account_name="Nexus Edutech Consult Ltd",
                bank_name="Zenith Bank Plc",
                account_number="1228490211",
                sort_code_or_swift="057150013",
                currency="NGN",
                payment_instructions="Please include your Institution Name and Invoice Reference in the transfer narration/remark.",
                support_email="billing@nexus.ng",
                is_active=True,
            )
        serializer = self.get_serializer(bank)
        return Response(serializer.data)


class PricingPlanViewSet(viewsets.ReadOnlyModelViewSet):
    """Provides public pricing tiers and subscription options."""

    queryset = PricingPlan.objects.filter(is_active=True).order_by("base_fee")
    serializer_class = PricingPlanSerializer
    permission_classes = [AllowAny]


class InstitutionInvoiceViewSet(viewsets.ModelViewSet):
    """Handles institution invoices, payment proof submissions, and verification."""

    queryset = InstitutionInvoice.objects.all().select_related("institution", "plan", "confirmed_by").order_by("-created_at")
    serializer_class = InstitutionInvoiceSerializer
    permission_classes = [AllowAny]
    lookup_field = "id"

    def get_queryset(self):
        qs = super().get_queryset()
        inst_id = self.request.query_params.get("institution_id")
        if inst_id:
            qs = qs.filter(institution_id=inst_id)
        elif self.request.user.is_authenticated and hasattr(self.request.user, "staff_profile") and self.request.user.staff_profile:
            qs = qs.filter(institution=self.request.user.staff_profile.institution)
        for inv in qs:
            if inv.status == InvoiceStatus.PAID and inv.institution.status != InstitutionStatus.ACTIVE:
                inv.institution.status = InstitutionStatus.ACTIVE
                inv.institution.save(update_fields=["status", "updated_at"])
        return qs

    @action(detail=True, methods=["post"], url_path="submit-payment")
    def submit_payment(self, request, id=None):
        import datetime
        invoice = self.get_object()
        serializer = InvoiceSubmitPaymentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data
        invoice.payment_reference = data.get("payment_reference", "").strip()
        invoice.payer_bank_name = data.get("payer_bank_name", "").strip()
        invoice.payer_account_name = data.get("payer_account_name", "").strip()
        invoice.payment_date = data.get("payment_date") or timezone.now().date()
        invoice.payment_notes = data.get("payment_notes", "").strip()

        if "payment_receipt_file" in request.FILES:
            invoice.payment_receipt_file = request.FILES["payment_receipt_file"]

        invoice.status = InvoiceStatus.PAYMENT_SUBMITTED
        invoice.payment_submitted_at = timezone.now()
        invoice.save()

        # Update institution status to PAYMENT_SUBMITTED
        institution = invoice.institution
        institution.status = InstitutionStatus.PAYMENT_SUBMITTED
        institution.save(update_fields=["status", "updated_at"])

        return Response(
            InstitutionInvoiceSerializer(invoice, context={"request": request}).data,
            status=status.HTTP_200_OK,
        )


class InstitutionRegistrationView(APIView):
    """
    Self-service institutional onboarding endpoint.
    Registers Institution, primary Admin User, creates initial divisions,
    and generates the onboarding invoice with snapshot company bank details.
    """

    permission_classes = [AllowAny]

    def post(self, request):
        import datetime
        import random
        from django.utils.text import slugify

        serializer = InstitutionRegistrationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        legal_name = data["legal_name"].strip()
        short_name = data["short_name"].strip()
        contact_email = data["contact_email"].lower().strip()
        contact_name = data["contact_name"].strip()
        password = data["password"]

        # Check existing user
        if User.objects.filter(email=contact_email).exists():
            return Response(
                {"error": f"A user with email '{contact_email}' already exists. Please sign in or use a different email."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check existing institution by name (case-insensitive) — name is unique
        if Institution.objects.filter(name__iexact=legal_name).exists():
            return duplicate_error(
                f"'{legal_name}' is already registered on the Nexus platform. "
                "If this is your institution, please contact support to claim your tenant."
            )

        # Reject duplicate faculty/school names before the unique-together insert
        faculties = data.get("faculties", [])
        clean_faculties = [f.strip() for f in faculties if f and f.strip()]
        seen_faculties: dict[str, str] = {}
        duplicate_faculties: list[str] = []
        for faculty_name in clean_faculties:
            key = faculty_name.casefold()
            if key in seen_faculties:
                duplicate_faculties.append(faculty_name)
            seen_faculties[key] = faculty_name
        if duplicate_faculties:
            return duplicate_error(
                "Duplicate division/faculty name(s) in the list: "
                + ", ".join(dict.fromkeys(duplicate_faculties))
                + ". Each faculty/school must have a unique name."
            )

        # Check existing institution
        base_slug = slugify(short_name) or slugify(legal_name) or f"inst-{random.randint(100, 999)}"
        slug = base_slug
        counter = 1
        while Institution.objects.filter(slug=slug).exists():
            slug = f"{base_slug}-{counter}"
            counter += 1

        with transaction.atomic():
            # 1. Create Institution
            institution = Institution.objects.create(
                name=legal_name,
                short_name=short_name,
                slug=slug,
                institution_type=data.get("institution_type", "UNIVERSITY"),
                ownership=data.get("ownership", "FEDERAL"),
                regulator=data.get("regulator", "NUC"),
                tier_two_term="FACULTY" if data.get("institution_type") == "UNIVERSITY" else "SCHOOL",
                state=data.get("state", "Niger"),
                address=data.get("address", ""),
                status=InstitutionStatus.PENDING_PAYMENT,
                domain_whitelist=[f"@{contact_email.split('@')[-1]}"] if "@" in contact_email else [],
            )

            # 2. Create Admin User
            user = User.objects.create_user(
                email=contact_email,
                name=contact_name,
                password=password,
            )

            # 3. Create InstitutionStaff & StaffAssignment (SUPERADMIN)
            InstitutionStaff.objects.create(
                user=user,
                institution=institution,
                role="SUPERADMIN",
                title=data.get("designation") or "Institutional Administrator",
                is_active=True,
            )
            StaffAssignment.objects.create(
                user=user,
                institution=institution,
                role_at_unit="SUPERADMIN",
                official_title=data.get("designation") or "Institutional Administrator",
                is_primary=True,
                is_active=True,
                can_evaluate_milestones=True,
                can_manage_waivers=True,
            )

            # 4. Create Scoped Divisions
            for idx, faculty_name in enumerate(clean_faculties, start=1):
                clean_name = faculty_name.strip()
                if clean_name:
                    code = "".join([w[0] for w in clean_name.split() if w]).upper()[:6] or f"FAC{idx}"
                    AcademicDivision.objects.create(
                        institution=institution,
                        name=clean_name,
                        code=code,
                        division_type=DivisionType.FACULTY if institution.institution_type == "UNIVERSITY" else DivisionType.SCHOOL,
                        is_active=True,
                    )

            # 5. Fetch Pricing Plan & Bank Details
            tier_code = data.get("tier", "standard").lower()
            plan = PricingPlan.objects.filter(code=tier_code, is_active=True).first()
            if not plan:
                plan = PricingPlan.objects.filter(is_active=True).first()
            if not plan:
                # Seed default plan
                plan = PricingPlan.objects.create(
                    code="standard",
                    name="Standard Tier",
                    target_institution_type="Universities & Polytechnics",
                    base_fee=1500000.00,
                    setup_onboarding_fee=150000.00,
                    max_students=8000,
                    currency="NGN",
                    features=[
                        "Sponsored onboarding & faculty calibration",
                        "Up to 8,000 student seats",
                        "All modules + analytics",
                        "Priority support & training",
                    ],
                )

            bank = CompanyBankDetail.objects.filter(is_active=True).first()
            if not bank:
                bank = CompanyBankDetail.objects.create(
                    account_name="Nexus Edutech Consult Ltd",
                    bank_name="Zenith Bank Plc",
                    account_number="1228490211",
                    sort_code_or_swift="057150013",
                    currency="NGN",
                    payment_instructions="Please include your Institution Name and Invoice Reference in the transfer narration/remark. Once transfer is completed, upload your payment receipt directly to the portal.",
                    support_email="billing@nexus.ng",
                    is_active=True,
                )

            # 6. Generate Invoice
            year = timezone.now().year
            seq = random.randint(1000, 9999)
            invoice_number = f"INV-NEXUS-{year}-{seq}"
            while InstitutionInvoice.objects.filter(invoice_number=invoice_number).exists():
                seq = random.randint(1000, 9999)
                invoice_number = f"INV-NEXUS-{year}-{seq}"

            from decimal import Decimal

            subtotal = plan.base_fee
            setup = plan.setup_onboarding_fee
            vat_rate = Decimal("7.50")
            taxable_amount = Decimal(str(subtotal)) + Decimal(str(setup))
            vat_amount = (taxable_amount * vat_rate / Decimal("100.00")).quantize(Decimal("0.01"))
            total = taxable_amount + vat_amount

            line_items = [
                {
                    "description": f"{plan.name} License Subscription (Annual)",
                    "quantity": 1,
                    "unit_price": float(subtotal),
                    "amount": float(subtotal),
                },
                {
                    "description": "Onboarding, Technical Tenant Provisioning & Faculty Calibration",
                    "quantity": 1,
                    "unit_price": float(setup),
                    "amount": float(setup),
                },
                {
                    "description": f"Value Added Tax (VAT @ {vat_rate}% Exclusive)",
                    "quantity": 1,
                    "unit_price": float(vat_amount),
                    "amount": float(vat_amount),
                },
            ]

            bank_snapshot = {
                "account_name": bank.account_name,
                "bank_name": bank.bank_name,
                "account_number": bank.account_number,
                "sort_code_or_swift": bank.sort_code_or_swift,
                "currency": bank.currency,
                "payment_instructions": bank.payment_instructions,
                "support_email": bank.support_email,
                "support_phone": bank.support_phone,
            }

            invoice = InstitutionInvoice.objects.create(
                invoice_number=invoice_number,
                institution=institution,
                plan=plan,
                plan_name=plan.name,
                issued_to_name=contact_name,
                issued_to_email=contact_email,
                subtotal_amount=subtotal,
                setup_fee=setup,
                vat_rate=vat_rate,
                vat_amount=vat_amount,
                discount_amount=0.00,
                total_amount=total,
                currency=plan.currency,
                status=InvoiceStatus.UNPAID,
                bank_details_snapshot=bank_snapshot,
                items_breakdown=line_items,
                due_date=timezone.now().date() + datetime.timedelta(days=30),
            )

            # Generate Token
            token, _ = Token.objects.get_or_create(user=user)

            return Response(
                {
                    "token": token.key,
                    "user": AuthUserSerializer(user).data,
                    "institution": InstitutionDetailSerializer(institution).data,
                    "invoice": InstitutionInvoiceSerializer(invoice, context={"request": request}).data,
                    "message": "Institution registered successfully. Please complete the bank transfer using the generated invoice.",
                },
                status=status.HTTP_201_CREATED,
            )






